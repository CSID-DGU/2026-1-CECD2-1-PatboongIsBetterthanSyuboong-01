"""상병/처방 코드 엑셀을 CSV로 변환한 뒤 MySQL 마스터 테이블에 적재한다.

기본 입력:
  - Back-End/상병코드.xlsx -> disease(code, name, name_en)
  - Back-End/처방코드.xlsx -> diagnose(code, name, dose, time, days)

Docker compose로 실행 중인 MySQL 컨테이너(bit-mysql)에 적재하는 것을 기본값으로 둔다.
"""
from __future__ import annotations

import argparse
import csv
import os
import subprocess
import sys
import tempfile
import unicodedata
from collections.abc import Iterable, Iterator
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_ROOT = SCRIPT_DIR.parent
DEFAULT_OUTPUT_DIR = BACKEND_ROOT / "generated" / "master-codes"
BATCH_SIZE = 5000

Target = Literal["disease", "diagnose"]


def find_input_file(preferred: Path, *fallback_names: str) -> Path:
    if preferred.is_file():
        return preferred

    available = {unicodedata.normalize("NFC", p.name): p for p in BACKEND_ROOT.iterdir() if p.is_file()}
    for name in fallback_names:
        hit = available.get(unicodedata.normalize("NFC", name))
        if hit and hit.is_file():
            return hit

    raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {preferred}")


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\x00", "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def iter_xlsx_rows(path: Path, *, min_columns: int) -> Iterator[list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        next(rows, None)  # header
        for row in rows:
            cells = [clean_cell(v) for v in row[:min_columns]]
            if len(cells) < min_columns:
                cells += [""] * (min_columns - len(cells))
            if not cells[0] or not cells[1]:
                continue
            yield cells
    finally:
        workbook.close()


def write_csv(path: Path, rows: Iterable[Iterable[str]], header: list[str]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in rows:
            writer.writerow(row)
            count += 1
    return count


def convert_disease_xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> int:
    def rows() -> Iterator[list[str]]:
        for code, name, name_en in iter_xlsx_rows(xlsx_path, min_columns=3):
            yield [code, name, "" if name_en == "0" else name_en]

    return write_csv(csv_path, rows(), ["code", "name", "name_en"])


def convert_diagnose_xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> int:
    def rows() -> Iterator[list[str]]:
        for code, name, *_ in iter_xlsx_rows(xlsx_path, min_columns=3):
            yield [code, name, "0", "0", "0"]

    return write_csv(csv_path, rows(), ["code", "name", "dose", "time", "days"])


def sql_str(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "''").replace("\x00", "") + "'"


def batched(rows: Iterable[list[str]], size: int) -> Iterator[list[list[str]]]:
    batch: list[list[str]] = []
    for row in rows:
        batch.append(row)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def iter_csv_rows(csv_path: Path) -> Iterator[list[str]]:
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if len(row) >= 2 and row[0].strip() and row[1].strip():
                yield [c.strip() for c in row]


def build_reset_sql(target: Target) -> str:
    table = "disease" if target == "disease" else "diagnose"
    create_sql = (
        """
CREATE TABLE IF NOT EXISTS disease (
  id INT NOT NULL AUTO_INCREMENT,
  code VARCHAR(128) NOT NULL,
  name TEXT NOT NULL,
  name_en TEXT NULL,
  PRIMARY KEY (id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
ALTER TABLE disease ADD COLUMN name_en TEXT NULL;
"""
        if target == "disease"
        else """
CREATE TABLE IF NOT EXISTS diagnose (
  id INT NOT NULL AUTO_INCREMENT,
  code VARCHAR(128) NOT NULL,
  name TEXT NOT NULL,
  dose INT NOT NULL DEFAULT 0,
  time INT NOT NULL DEFAULT 0,
  days INT NOT NULL DEFAULT 0,
  PRIMARY KEY (id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
ALTER TABLE diagnose MODIFY COLUMN name TEXT NOT NULL;
"""
    )
    return f"""SET NAMES utf8mb4;
{create_sql}
SET FOREIGN_KEY_CHECKS = 0;
DELETE FROM {table};
ALTER TABLE {table} AUTO_INCREMENT = 1;
SET FOREIGN_KEY_CHECKS = 1;
"""


def build_insert_sql(target: Target, batch: list[list[str]]) -> str:
    if target == "disease":
        values = []
        for code, name, name_en, *_ in batch:
            ne = "NULL" if not name_en else sql_str(name_en)
            values.append(f"({sql_str(code)},{sql_str(name)},{ne})")
        return "SET NAMES utf8mb4;\nINSERT INTO disease (code, name, name_en) VALUES " + ",".join(values) + ";"

    values = []
    for code, name, dose, time, days, *_ in batch:
        values.append(f"({sql_str(code)},{sql_str(name)},{int_or_zero(dose)},{int_or_zero(time)},{int_or_zero(days)})")
    return "SET NAMES utf8mb4;\nINSERT INTO diagnose (code, name, dose, time, days) VALUES " + ",".join(values) + ";"


def int_or_zero(value: str) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


class MysqlRunner:
    def __init__(self, args: argparse.Namespace) -> None:
        self.mode = args.mysql_mode
        self.container = args.mysql_container
        self.database = args.mysql_database
        self.host = args.mysql_host
        self.port = str(args.mysql_port)
        self.user = args.mysql_user
        self.password = args.mysql_password
        self._cnf_path: str | None = None

    def __enter__(self) -> "MysqlRunner":
        if self.mode == "local":
            cnf = tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".cnf")
            cnf.write(
                f"""[client]
host={self.host}
port={self.port}
user={self.user}
password={self.password}
default-character-set=utf8mb4
"""
            )
            cnf.flush()
            self._cnf_path = cnf.name
            cnf.close()
        return self

    def __exit__(self, *_: object) -> None:
        if self._cnf_path:
            Path(self._cnf_path).unlink(missing_ok=True)

    def execute(self, sql: str, *, tolerate_duplicate_column: bool = False) -> None:
        if self.mode == "docker":
            cmd = [
                "docker",
                "exec",
                "-i",
                self.container,
                "mysql",
                "--default-character-set=utf8mb4",
                f"-u{self.user}",
                f"-p{self.password}",
                self.database,
            ]
        else:
            if not self._cnf_path:
                raise RuntimeError("local mysql config was not initialized")
            cmd = ["mysql", f"--defaults-extra-file={self._cnf_path}", self.database]

        result = subprocess.run(cmd, input=sql, capture_output=True, text=True, encoding="utf-8")
        if result.returncode != 0:
            err = result.stderr or result.stdout
            if tolerate_duplicate_column and ("Duplicate column" in err or "1060" in err):
                return
            raise RuntimeError(err.strip())


def import_csv(target: Target, csv_path: Path, runner: MysqlRunner) -> int:
    reset_sql = build_reset_sql(target)
    if target == "disease":
        create_sql, purge_sql = reset_sql.split("ALTER TABLE disease ADD COLUMN name_en TEXT NULL;")
        runner.execute(create_sql)
        runner.execute("ALTER TABLE disease ADD COLUMN name_en TEXT NULL;", tolerate_duplicate_column=True)
        runner.execute(purge_sql)
    else:
        runner.execute(reset_sql)

    total = 0
    for batch in batched(iter_csv_rows(csv_path), BATCH_SIZE):
        runner.execute(build_insert_sql(target, batch))
        total += len(batch)
    return total


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--target", choices=["all", "disease", "diagnose"], default="all")
    parser.add_argument("--disease-xlsx", type=Path, default=BACKEND_ROOT / "상병코드.xlsx")
    parser.add_argument("--diagnose-xlsx", type=Path, default=BACKEND_ROOT / "처방코드.xlsx")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--convert-only", action="store_true", help="CSV만 생성하고 DB 적재는 하지 않습니다.")
    parser.add_argument("--mysql-mode", choices=["docker", "local"], default=os.environ.get("MYSQL_MODE", "docker"))
    parser.add_argument("--mysql-container", default=os.environ.get("MYSQL_CONTAINER", "bit-mysql"))
    parser.add_argument("--mysql-host", default=os.environ.get("MYSQL_HOST", "127.0.0.1"))
    parser.add_argument("--mysql-port", default=os.environ.get("MYSQL_PORT", "3307"))
    parser.add_argument("--mysql-user", default=os.environ.get("MYSQL_USER", "root"))
    parser.add_argument("--mysql-password", default=os.environ.get("MYSQL_PASSWORD", "Dd905925@"))
    parser.add_argument("--mysql-database", default=os.environ.get("MYSQL_DATABASE", "bitcomputer"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    targets: list[Target] = ["disease", "diagnose"] if args.target == "all" else [args.target]
    outputs: dict[Target, Path] = {}

    if "disease" in targets:
        disease_xlsx = find_input_file(args.disease_xlsx, "상병코드.xlsx", "상병코드.xlsx")
        disease_csv = args.output_dir / "disease_codes.csv"
        count = convert_disease_xlsx_to_csv(disease_xlsx, disease_csv)
        outputs["disease"] = disease_csv
        print(f"[convert] disease: {count} rows -> {disease_csv}")

    if "diagnose" in targets:
        diagnose_xlsx = find_input_file(args.diagnose_xlsx, "처방코드.xlsx", "처방코드.xlsx")
        diagnose_csv = args.output_dir / "prescription_codes.csv"
        count = convert_diagnose_xlsx_to_csv(diagnose_xlsx, diagnose_csv)
        outputs["diagnose"] = diagnose_csv
        print(f"[convert] diagnose: {count} rows -> {diagnose_csv}")

    if args.convert_only:
        return 0

    with MysqlRunner(args) as runner:
        for target in targets:
            imported = import_csv(target, outputs[target], runner)
            table = "disease" if target == "disease" else "diagnose"
            print(f"[import] {table}: {imported} rows")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
